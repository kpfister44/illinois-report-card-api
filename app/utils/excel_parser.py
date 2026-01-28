# ABOUTME: Excel file parsing utilities
# ABOUTME: Reads Report Card Excel files and extracts data from multiple sheets

import openpyxl
from typing import Dict, List, Any, Optional


def parse_excel_file(
    file_path: str,
    sheets: Optional[List[str]] = None,
    max_rows: Optional[int] = None
) -> Dict[str, Dict[str, Any]]:
    """
    Parse an Excel file and return structured data from specified sheets.

    Args:
        file_path: Path to Excel file
        sheets: List of sheet names to parse (None = all sheets)
        max_rows: Maximum number of data rows to read per sheet (None = all rows)

    Returns:
        Dict mapping sheet names to their data:
        {
            "General": {
                "headers": ["RCDTS", "School Name", ...],
                "rows": [
                    {"RCDTS": "...", "School Name": "...", ...},
                    ...
                ]
            },
            ...
        }
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Determine which sheets to parse
    if sheets is None:
        sheets_to_parse = wb.sheetnames
    else:
        sheets_to_parse = [s for s in sheets if s in wb.sheetnames]

    result = {}

    for sheet_name in sheets_to_parse:
        sheet = wb[sheet_name]

        # Get header row (first row)
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(header_row)]

        # Filter out None headers at the end
        # Find last non-None header
        last_valid_idx = len(headers) - 1
        for i in range(len(headers) - 1, -1, -1):
            if not headers[i].startswith('Column_'):
                last_valid_idx = i
                break

        headers = headers[:last_valid_idx + 1]

        # Get data rows
        rows = []
        row_iterator = sheet.iter_rows(min_row=2, values_only=True)

        if max_rows:
            row_iterator = list(row_iterator)[:max_rows]

        for row_data in row_iterator:
            # Convert row tuple to dict, handling None values
            row_dict = {}
            for i, header in enumerate(headers):
                value = row_data[i] if i < len(row_data) else None

                # Convert empty strings to None
                if value == "":
                    value = None

                row_dict[header] = value

            rows.append(row_dict)

        result[sheet_name] = {
            "headers": headers,
            "rows": rows
        }

    wb.close()
    return result

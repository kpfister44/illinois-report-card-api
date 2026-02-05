# ABOUTME: Pytest fixtures for CLI import_data module unit tests
# ABOUTME: Provides Excel test files and database fixtures for comprehensive CLI testing

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook
from sqlalchemy import create_engine
from app.models.database import Base


@pytest.fixture
def empty_excel_file(tmp_path):
    """Create an empty Excel file with no sheets for testing error handling."""
    file_path = tmp_path / "empty.xlsx"
    # Create workbook but don't add any sheets - this tests the empty case
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def excel_missing_general_sheet(tmp_path):
    """Create an Excel file with sheets but no 'General' sheet."""
    file_path = tmp_path / "missing_general.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Finance"  # Wrong sheet name
    ws.append(["Column1", "Column2"])
    ws.append(["Value1", "Value2"])
    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def excel_empty_general_sheet(tmp_path):
    """Create an Excel file with 'General' sheet but no data rows (only headers)."""
    file_path = tmp_path / "empty_general.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "General"
    # Add headers but no data rows
    ws.append(["RCDTS", "School Name", "City", "County"])
    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def excel_with_edge_cases(tmp_path):
    """
    Create an Excel file with edge case data: NULL values, suppressed data (*),
    commas in numbers, percentages, etc.
    """
    file_path = tmp_path / "edge_cases.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "General"

    # Headers
    headers = [
        "RCDTS", "School Name", "City", "County", "Type",
        "Enrollment", "Pct Low Income", "SAT Average", "Graduation Rate %"
    ]
    ws.append(headers)

    # Row 1: Normal data
    ws.append([
        "01-001-0010-26-0001",
        "Test Elementary School",
        "Springfield",
        "Sangamon",
        "School",
        "425",
        "45.5%",
        "980",
        "92.3%"
    ])

    # Row 2: Edge cases - commas in enrollment, suppressed SAT, high percentage
    ws.append([
        "01-001-0010-26-0002",
        "Test High School",
        "Chicago",
        "Cook",
        "School",
        "1,250",  # Comma in number
        "38.2%",
        "*",  # Suppressed value
        "95.1%"
    ])

    # Row 3: More suppressed values and NULL-like data
    ws.append([
        "01-001-0010-26-0003",
        "Test Middle School",
        "Peoria",
        "Peoria",
        "School",
        "650",
        "<10",  # Suppressed (less than 10)
        None,  # NULL value
        "88.5%"
    ])

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def excel_without_rcdts(tmp_path):
    """Create an Excel file with valid data but no RCDTS column."""
    file_path = tmp_path / "no_rcdts.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "General"

    # Headers without RCDTS
    headers = ["School Name", "City", "County", "Enrollment"]
    ws.append(headers)

    # Data rows
    ws.append(["Test School 1", "Springfield", "Sangamon", "425"])
    ws.append(["Test School 2", "Chicago", "Cook", "850"])

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def test_excel_file(tmp_path):
    """
    Create a standard test Excel file with realistic school data.
    Reusable fixture for happy path tests.
    """
    file_path = tmp_path / "test_data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "General"

    # Headers
    headers = [
        "RCDTS", "School Name", "City", "County", "Type",
        "Enrollment", "Pct Low Income", "SAT Average", "Graduation Rate %"
    ]
    ws.append(headers)

    # Sample data rows
    ws.append([
        "01-001-0010-26-0001",
        "Test Elementary School",
        "Springfield",
        "Sangamon",
        "School",
        "425",
        "45.5%",
        "980",
        "92.3%"
    ])

    ws.append([
        "01-001-0010-26-0002",
        "Test High School",
        "Springfield",
        "Sangamon",
        "School",
        "1,250",
        "38.2%",
        "1150",
        "95.1%"
    ])

    ws.append([
        "01-001-0010-26-0003",
        "Test Middle School",
        "Chicago",
        "Cook",
        "School",
        "650",
        "55.0%",
        "*",
        "88.5%"
    ])

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def temp_database(tmp_path):
    """
    Create a temporary SQLite database file for testing.
    Unlike in-memory databases, this allows testing engine disposal.
    """
    db_path = tmp_path / "test_import.db"
    db_url = f"sqlite:///{db_path}"

    # Create engine and initialize schema
    engine = create_engine(db_url)
    Base.metadata.create_all(bind=engine)

    # Setup FTS5 for full-text search
    from app.services.fts5 import setup_fts5
    setup_fts5(engine)

    engine.dispose()

    yield str(db_path)

    # Cleanup
    if db_path.exists():
        db_path.unlink()
